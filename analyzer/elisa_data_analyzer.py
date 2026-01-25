import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats # Kept for stats.sem in plotting
# core imported below
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import elisa_core

class AnalysisConfigDialog:
    def __init__(self, parent, timepoints):
        self.top = tk.Toplevel(parent)
        self.top.title("Analysis Options")
        self.result = None
        self.timepoints = sorted(list(timepoints))
        
        tk.Label(self.top, text="Select Timepoints to Compare:", font=('Arial', 10, 'bold')).pack(pady=5)
        
        self.vars = {}
        frame_tp = tk.Frame(self.top)
        frame_tp.pack(pady=5, padx=10)
        
        for tp in self.timepoints:
            var = tk.BooleanVar(value=True)
            self.vars[tp] = var
            tk.Checkbutton(frame_tp, text=tp, variable=var).pack(anchor='w')            
        tk.Label(self.top, text="Configuration:", font=('Arial', 10, 'bold')).pack(pady=5)
        
        # Paired
        self.paired_var = tk.BooleanVar(value=True)
        tk.Checkbutton(self.top, text="Paired / Repeated Measures", variable=self.paired_var).pack(anchor='w', padx=20)
        
        # Post Hoc
        self.posthoc_var = tk.BooleanVar(value=False)
        tk.Checkbutton(self.top, text="Perform Post-Hoc (if significant)", variable=self.posthoc_var).pack(anchor='w', padx=20)
        
        # Tails
        tk.Label(self.top, text="Alternative Hypothesis (T-Test):").pack(pady=(5, 0))
        self.tail_var = tk.StringVar(value="two-sided")
        ttk.Combobox(self.top, textvariable=self.tail_var, values=["two-sided", "less", "greater"], state="readonly").pack(pady=2)
        tk.Label(self.top, text="less: t0 is lower | greater: t0 is higher", font=('Arial', 8)).pack(pady=1)
        
        btn = tk.Button(self.top, text="Run Analysis", command=self.on_submit, bg="#ccffcc")
        btn.pack(pady=10)
        
        # Force visibility
        self.top.lift()
        self.top.attributes('-topmost', True)
        self.top.grab_set()
        parent.wait_window(self.top)
        
    def on_submit(self):
        selected = [tp for tp, var in self.vars.items() if var.get()]
        if len(selected) < 2:
            messagebox.showwarning("Warning", "Please select at least 2 timepoints.")
            return
        
        self.result = {
            'timepoints': selected,
            'paired': self.paired_var.get(),
            'tails': self.tail_var.get(),
            'posthoc': self.posthoc_var.get()
        }
        self.top.destroy()

class ElisaAnalyzer:
    def __init__(self):
        self.layout_df = None
        self.cal_means = None
        self.od450_df = None
        self.od630_df = None
        self.merged_df = None
        self.analyzed_df = None
        self.calibration_model = None
        self.stats_results = {}
        self.layout_path = ""
        self.instrument_path = ""
        self.config = None
        self.root = tk.Tk()
        self.root.withdraw()

    def __del__(self):
        try:
            self.root.destroy()
        except:
            pass

    def load_files(self):
        """Opens file dialogs for Layout CSV and Instrument Excel."""
        # 1. Select Layout CSV
        messagebox.showinfo("Select File", "Please select the Layout CSV file.")
        layout_path = filedialog.askopenfilename(title="Select Layout CSV", filetypes=[("CSV files", "*.csv")])
        if not layout_path: return False
        
        # 2. Select Instrument Excel
        messagebox.showinfo("Select File", "Please select the Tecan Instrument Excel export.")
        instrument_path = filedialog.askopenfilename(title="Select Tecan Excel Output", filetypes=[("Excel files", "*.xlsx;*.xls")])
        if not instrument_path: return False

        self.layout_path = layout_path
        self.instrument_path = instrument_path
        
        try:
            self.layout_df = pd.read_csv(layout_path)
            self.parse_instrument_excel(instrument_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load files:\n{e}")
            return False

    def parse_instrument_excel(self, path):
        """Robustly parses Tecan Excel using core logic."""
        try:
            self.od450_df, self.od630_df = elisa_core.parse_tecan_excel(path)
            print("Successfully extracted OD450 and OD630 grids.")
        except Exception as e:
            raise e

    # extract_grid removed (moved to core)

    def process_data(self):
        """Merges data, corrects OD, and runs calibration using core logic."""
        # 1. Merge & Correct
        self.merged_df = elisa_core.merge_and_correct(self.layout_df, self.od450_df, self.od630_df)
        
        # 2. Calibration
        self.calibration_model, self.cal_means = elisa_core.fit_calibration_model(self.merged_df)
        
        if self.calibration_model:
            print(f"Calibration: OD = {self.calibration_model['slope']:.4f} * Conc + {self.calibration_model['intercept']:.4f} (R2={self.calibration_model['r_squared']:.4f})")
            
        # 3. Calculate Concentrations
        self.analyzed_df = elisa_core.calculate_concentrations(self.merged_df, self.calibration_model)
        
        return self.analyzed_df

    def configure_analysis(self):
        """Opens dialog to configure analysis."""
        exp_df = self.analyzed_df[self.analyzed_df['Type'] == 'Experiment']
        if exp_df.empty: return False
        
        # Get Timepoints
        timepoints = exp_df['Timepoint'].unique()
        
        self.root.update()
        dlg = AnalysisConfigDialog(self.root, timepoints)
        
        if dlg.result:
            self.config = dlg.result
            return True
        return False

    def run_calibration(self, df):
        # Deprecated: Logic moved to process_data via elisa_core
        return df

    def run_statistics(self):
        """Performs statistical analysis using core logic."""
        exp_df = self.analyzed_df[self.analyzed_df['Type'] == 'Experiment'].copy()
        
        if exp_df.empty:
            print("No experiment data found.")
            return

        # Run Analysis
        self.stats_results = elisa_core.run_statistical_analysis(exp_df, self.config)
        
        if 'test_decision' in self.stats_results:
            print(f"\nStats Decision: {self.stats_results['test_decision']}, p={self.stats_results['p_value']:.5f}")
            


    def generate_plots(self):
        """Generates Calibration and Result plots."""
        # 1. Calibration Curve
        cal_data = self.merged_df[self.merged_df['Type'] == 'Calibration']
        if not cal_data.empty:
            plt.figure(figsize=(6, 4))
            sns.regplot(x='Concentration', y='OD_Corr', data=cal_data, ci=None, label='Standards')
            
            # Add Model Line
            x_range = np.linspace(cal_data['Concentration'].min(), cal_data['Concentration'].max(), 100)
            y_model = self.calibration_model['slope'] * x_range + self.calibration_model['intercept']
            plt.plot(x_range, y_model, 'r--', label=f"R2={self.calibration_model['r_squared']:.4f}")
            
            plt.title("Calibration Curve")
            plt.xlabel("Insulin (ng/mL)")
            plt.ylabel("Absorbance")
            plt.legend()
            plt.tight_layout()
            plt.savefig("calibration_curve.png")
            plt.close()
            
        # 2. Results Bar Graph
        exp_df = self.analyzed_df[self.analyzed_df['Type'] == 'Experiment']
        if not exp_df.empty:
            plt.figure(figsize=(10, 6))
            # Plot Mean Concentration per Subject/Timepoint
            sns.barplot(x='Subject Name', y='Calculated_Conc', hue='Timepoint', data=exp_df, errorbar='se', capsize=.1)
            
            plt.title("Insulin concentration in time")
            plt.ylabel("Insulin (ng/mL)")
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig("results_bar_graph.png")
            plt.close()
            
            # 3. Significance Plot (Aggregated by Timepoint)
            plt.figure(figsize=(6, 5))
            
            # Filter to selected timepoints if config exists
            plot_df = exp_df
            if self.config:
                plot_df = exp_df[exp_df['Timepoint'].isin(self.config['timepoints'])]
            
            # Plot
            # Fix Future Warning: Assign x to hue
            ax = sns.barplot(x='Timepoint', y='Calculated_Conc', hue='Timepoint', data=plot_df, errorbar='se', capsize=.1, palette="pastel", legend=False)
            plt.title("Mean Insulin")
            plt.ylabel("Insulin (ng/mL)")
            
            # Add Significance Bars
            if 'p_value' in self.stats_results and self.stats_results['p_value'] is not np.nan:
                p_val = self.stats_results['p_value']
                unique_tps = sorted(plot_df['Timepoint'].unique())
                
                # Only if 2 groups for now
                if len(unique_tps) == 2:
                    # Get y-max for bracket placement
                    y_max = plot_df.groupby('Timepoint')['Calculated_Conc'].mean().max() + plot_df.groupby('Timepoint')['Calculated_Conc'].apply(stats.sem).max()
                    y_h = y_max * 1.05
                    y_h2 = y_max * 1.10
                    
                    # Sig Stars
                    star = "ns"
                    if p_val <= 0.001: star = "***"
                    elif p_val <= 0.01: star = "**"
                    elif p_val <= 0.05: star = "*"
                    
                    # Draw Line
                    x1, x2 = 0, 1
                    plt.plot([x1, x1, x2, x2], [y_h, y_h2, y_h2, y_h], lw=1.5, c='k')
                    plt.text((x1+x2)*.5, y_h2, star, ha='center', va='bottom', color='k', fontsize=12, fontweight='bold')
            
            plt.tight_layout()
            plt.savefig("average_plot.png")
            plt.close()



    def save_results(self):
        """Saves results as Pivot Tables to Instrument File."""
        try:
            wb = load_workbook(self.instrument_path)
            # Create unique sheet name
            base_name = "Analysis_Res"
            count = 1
            sheet_name = base_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name}_{count}"
                count += 1
            
            ws = wb.create_sheet(sheet_name)
            
            self._write_analysis_to_sheet(ws, start_row=1)

            wb.save(self.instrument_path)
            print(f"Analysis saved to sheet '{sheet_name}' in: {self.instrument_path}")
            
        except PermissionError:
            messagebox.showerror("Error", "Permission Denied! Close the Excel file and try again.")
            return
        except Exception as e:
            print(f"Detailed Error: {e}")
            messagebox.showerror("Error", f"Failed to save to Excel: {e}")
            return

    def save_to_master(self):
        """Saves Raw Data + Analysis to a Master File."""
        # 1. Ask for Master File
        messagebox.showinfo("Master File", "Please select the Master Excel File to append results to.")
        master_path = filedialog.askopenfilename(title="Select Master File", filetypes=[("Excel files", "*.xlsx")])
        if not master_path: return
        
        # 2. Ask for Sheet Name (Loop for Validation)
        while True:
            sheet_name = simpledialog.askstring("Sheet Name", "Enter a name for the new sheet:\n(Max 31 chars, no \\ / ? * [ ] :)", parent=self.root)
            if not sheet_name: return # Cancelled
            
            # Excel Invalid Chars Validation
            invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
            if any(char in sheet_name for char in invalid_chars):
                messagebox.showerror("Invalid Name", "Sheet name cannot contain characters: \\ / ? * [ ] :")
                continue
            if len(sheet_name) > 31:
                messagebox.showerror("Invalid Name", "Sheet name must be 31 characters or less.")
                continue
                
            break
        
        try:
            try:
                wb = load_workbook(master_path)
            except FileNotFoundError:
                from openpyxl import Workbook
                wb = Workbook()
                # If new, remove default sheet
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
            
            # Check overlap
            if sheet_name in wb.sheetnames:
                messagebox.showwarning("Warning", f"Sheet '{sheet_name}' already exists. A number will be appended.")
                count = 1
                base = sheet_name
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{base}_{count}"
                    count += 1
            
            ws = wb.create_sheet(sheet_name)
            current_row = 1
            
            # --- WRITE TECAN DATA ---
            try:
                wb_instr = load_workbook(self.instrument_path)
                ws_instr = wb_instr.active # Assuming first sheet
                
                # Copy all rows
                for row in ws_instr.iter_rows(values_only=True):
                    ws.append(row)
                
                current_row = ws.max_row + 4 # Add gap
            except Exception as e:
                print(f"Error copying Tecan data: {e}")
                ws.cell(row=current_row, column=1, value="Error copying Tecan data")
                current_row += 2
            
            # --- WRITE ANALYSIS ---
            self._write_analysis_to_sheet(ws, start_row=current_row)
            
            wb.save(master_path)
            messagebox.showinfo("Success", f"Saved to Master File:\n{master_path}\nSheet: {sheet_name}")
            
        except Exception as e:
             messagebox.showerror("Error", f"Failed to save to Master File:\n{e}")

    def _write_analysis_to_sheet(self, ws, start_row=1):
        """Writes the standard analysis tables to the given worksheet starting at start_row."""
        current_row = start_row
        
        # --- SECTION 0: Summary Metrics (New) ---
        ws.cell(row=current_row, column=1, value="Metric")
        ws.cell(row=current_row, column=2, value="Value")
        ws.cell(row=current_row, column=3, value="Test")
        
        if 'p_value' in self.stats_results:
            ws.cell(row=current_row+1, column=1, value="P-Value")
            ws.cell(row=current_row+1, column=2, value=self.stats_results['p_value'])
            ws.cell(row=current_row+1, column=3, value=self.stats_results.get('test_decision', 'N/A'))
        
        current_row += 4
        
        # --- SECTION 1: Calibration Equation ---
        ws.cell(row=current_row, column=1, value="Calibration Curve")
        if self.calibration_model:
            eq = f"y = {self.calibration_model['slope']:.4f}x + {self.calibration_model['intercept']:.4f}"
            r2 = f"R2 = {self.calibration_model['r_squared']:.4f}"
            ws.cell(row=current_row+1, column=1, value=eq)
            ws.cell(row=current_row+2, column=1, value=r2)
        else:
            ws.cell(row=current_row+1, column=1, value="No Calibration Data")
        
        current_row += 4

        # --- SECTION 2: Calibration Data Table ---
        if self.cal_means is not None:
            ws.cell(row=current_row, column=1, value="Calibration Data")
            ws.cell(row=current_row+1, column=1, value="Insulin (ng/mL)")
            ws.cell(row=current_row+1, column=2, value="Mean Abs")
            
            cal_rows = list(dataframe_to_rows(self.cal_means, index=False, header=False))
            for r_idx, row in enumerate(cal_rows, 1):
                ws.cell(row=current_row+1+r_idx, column=1, value=row[0])
                ws.cell(row=current_row+1+r_idx, column=2, value=row[1])
            
            current_row += len(cal_rows) + 3

        # --- PREPARE EXPERIMENT DATA ---
        exp_df = self.analyzed_df[self.analyzed_df['Type'] == 'Experiment'].copy()
        if not exp_df.empty:
            # Pivot 1: Mean Absorbance (OD_Corr)
            mean_abs = exp_df.groupby(['Subject Name', 'Timepoint'])['OD_Corr'].mean().reset_index()
            pivot_abs = mean_abs.pivot(index='Subject Name', columns='Timepoint', values='OD_Corr')
            
            # Pivot 2: Mean Concentration
            mean_conc = exp_df.groupby(['Subject Name', 'Timepoint'])['Calculated_Conc'].mean().reset_index()
            pivot_conc = mean_conc.pivot(index='Subject Name', columns='Timepoint', values='Calculated_Conc')

            # --- SECTION 3: Mean Absorbance Table ---
            ws.cell(row=current_row, column=1, value="Mean Absorbance")
            
            # Write Header (Timepoints)
            for c_idx, col_name in enumerate(pivot_abs.columns, 2):
                ws.cell(row=current_row+1, column=c_idx, value=col_name)
            
            # Write Rows (Subject + Values)
            for r_idx, (subject, row) in enumerate(pivot_abs.iterrows(), 1):
                ws.cell(row=current_row+1+r_idx, column=1, value=subject)
                for c_idx, val in enumerate(row, 2):
                    ws.cell(row=current_row+1+r_idx, column=c_idx, value=val)
            
            current_row += len(pivot_abs) + 3

            # --- SECTION 4: Mean Concentration Table ---
            ws.cell(row=current_row, column=1, value="Mean Insulin (ng/mL)")
            
            # Write Header
            for c_idx, col_name in enumerate(pivot_conc.columns, 2):
                ws.cell(row=current_row+1, column=c_idx, value=col_name)
            
            # Write Rows
            for r_idx, (subject, row) in enumerate(pivot_conc.iterrows(), 1):
                ws.cell(row=current_row+1+r_idx, column=1, value=subject)
                for c_idx, val in enumerate(row, 2):
                    ws.cell(row=current_row+1+r_idx, column=c_idx, value=val)
            
            current_row += len(pivot_conc) + 3

        # --- SECTION 5: Statistics ---
        ws.cell(row=current_row, column=1, value="Statistical Analysis (t0 vs t1)") # Label slightly hardcoded but acceptable
        
        if 'p_value' in self.stats_results:
            # Dynamic Shapiro Output
            r_offset = 1
            if 'shapiro' in self.stats_results:
                for tp, p_val in self.stats_results['shapiro'].items():
                        ws.cell(row=current_row + r_offset, column=1, value=f"Normality (Shapiro) {tp}: p={p_val:.4f}")
                        r_offset += 1
            
            ws.cell(row=current_row + r_offset, column=1, value=f"Homogeneity (Levene): p={self.stats_results.get('p_levene', 'N/A'):.4f}")
            r_offset += 2
            
            ws.cell(row=current_row + r_offset, column=1, value=f"Test: {self.stats_results.get('test_decision', 'N/A')}")
            ws.cell(row=current_row + r_offset + 1, column=1, value=f"P-Value: {self.stats_results['p_value']}")
            
            if 'posthoc' in self.stats_results:
                    ws.cell(row=current_row + r_offset + 2, column=1, value="Post-Hoc Analysis (Bonferroni)")
                    ph_df = self.stats_results['posthoc']
                    ph_rows = dataframe_to_rows(ph_df, index=False, header=True)
                    for r, row in enumerate(ph_rows):
                        for c, val in enumerate(row):
                            ws.cell(row=current_row + r_offset + 3 + r, column=c + 1, value=val)
                    r_offset += len(ph_df) + 4
            else:
                    r_offset += 2
            
            if 'High_CV' in self.stats_results:
                    ws.cell(row=current_row + r_offset + 1, column=1, value="High CV Warnings (>20%)")
                    high_cv = self.stats_results['High_CV']
                    # Write CV table
                    cv_rows = dataframe_to_rows(high_cv, index=False, header=True)
                    for r, row in enumerate(cv_rows):
                        for c, val in enumerate(row):
                            ws.cell(row=current_row + r_offset + 2 + r, column=c + 1, value=val)

    def run(self):
        """
        Main execution flow of the application.
        
        Steps:
        1. Load Layout and Instrument files.
        2. Process data (merge, correct, calibrate).
        3. Configure analysis options via GUI.
        4. Run statistical tests.
        5. Generate plots.
        6. Save results to Instrument file.
        7. Export to Master file.
        """
        if self.load_files():
            self.process_data()
            if self.configure_analysis(): 
                self.run_statistics()
                self.generate_plots()
                self.save_results()
                self.save_to_master()
                messagebox.showinfo("Success", "Analysis Complete!")


if __name__ == "__main__":
    app = ElisaAnalyzer()
    app.run()
