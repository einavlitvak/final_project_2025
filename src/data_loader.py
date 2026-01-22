import openpyxl
import pandas as pd
import numpy as np

class DataReader:
    """
    Reads ELISA plate reader output (Excel).
    """
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.sheet = self.wb.active

    def extract_data_grid(self):
        """
        Returns: pd.DataFrame (8 rows x 12 columns)
        """
        start_row = self._find_start_row()
        if start_row is None:
            raise ValueError("Could not find start of plate data.")
        
        data = []
        rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        
        for i, r_label in enumerate(rows):
            current_row_idx = start_row + i
            row_data = []
            for c_idx in range(1, 13):
                excel_col = c_idx + 1 # Block starts at Col 1 (Label), data at Col 2
                val = self.sheet.cell(row=current_row_idx, column=excel_col).value
                try:
                    val = float(val)
                except:
                    val = np.nan
                row_data.append(val)
            data.append(row_data)

        df = pd.DataFrame(data, index=rows, columns=list(range(1, 13)))
        return df

    def _find_start_row(self):
        for r in range(1, 100):
            val = self.sheet.cell(row=r, column=1).value
            if str(val).strip() == 'A':
                val_next = self.sheet.cell(row=r+1, column=1).value
                if str(val_next).strip() == 'B':
                    return r
        return None
